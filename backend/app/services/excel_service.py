"""Excel 가져오기/내보내기 서비스.

openpyxl을 사용하여 리마인더 데이터를 Excel 파일로 내보내거나
Excel 파일에서 리마인더 데이터를 가져옵니다.
"""
import io
from datetime import date, datetime
from uuid import UUID
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from fastapi import HTTPException, status, UploadFile
from app.models.reminder import Reminder
from app.models.company import CompanyMember

CATEGORY_MAP = {
    "원천세": "원천세",
    "4대보험": "4대보험",
    "부가세": "부가세",
    "법인세": "법인세",
    "종합소득세": "종합소득세",
    "지방소득세": "지방소득세",
    "연말정산": "연말정산",
    "급여": "급여",
    "HR": "HR",
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="맑은 고딕", size=10)
COMPLETED_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
OVERDUE_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


async def export_reminders_to_excel(
    db: AsyncSession,
    company_id: UUID,
    user_id: UUID,
    year: int | None = None,
    category: str | None = None,
) -> io.BytesIO:
    """리마인더를 Excel 파일로 내보냅니다."""
    # 접근 권한 확인
    member_result = await db.execute(
        select(CompanyMember).where(
            CompanyMember.user_id == user_id,
            CompanyMember.company_id == company_id,
        )
    )
    if not member_result.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied",
        )

    query = select(Reminder).where(Reminder.company_id == company_id)
    if year:
        from sqlalchemy import func
        query = query.where(func.extract("year", Reminder.deadline) == year)
    if category:
        query = query.where(Reminder.category == category)
    query = query.order_by(Reminder.deadline)

    result = await db.execute(query)
    reminders = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "일정 목록"

    # 헤더
    headers = ["번호", "제목", "카테고리", "마감일", "원래 마감일", "완료 여부", "우선순위", "설명"]
    col_widths = [8, 40, 15, 15, 15, 12, 10, 50]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[chr(64 + col_idx)].width = width

    today = date.today()

    for row_idx, reminder in enumerate(reminders, 2):
        row_data = [
            row_idx - 1,
            reminder.title,
            reminder.category,
            reminder.deadline.strftime("%Y-%m-%d"),
            reminder.original_deadline.strftime("%Y-%m-%d") if reminder.original_deadline else "",
            "완료" if reminder.completed else "미완료",
            _priority_label(reminder.priority),
            reminder.description or "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            if col_idx in (1, 4, 5, 6, 7):
                cell.alignment = Alignment(horizontal="center")

        if reminder.completed:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = COMPLETED_FILL
        elif reminder.deadline < today and not reminder.completed:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = OVERDUE_FILL

    # 요약 시트
    _add_summary_sheet(wb, reminders)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _priority_label(priority: int) -> str:
    labels = {0: "보통", 1: "낮음", 2: "높음", 3: "긴급"}
    return labels.get(priority, "보통")


def _add_summary_sheet(wb: Workbook, reminders: list) -> None:
    """요약 시트를 추가합니다."""
    ws = wb.create_sheet("요약")

    ws.cell(row=1, column=1, value="카테고리별 요약").font = Font(size=14, bold=True)

    categories = {}
    for r in reminders:
        cat = r.category
        if cat not in categories:
            categories[cat] = {"total": 0, "completed": 0}
        categories[cat]["total"] += 1
        if r.completed:
            categories[cat]["completed"] += 1

    headers = ["카테고리", "전체", "완료", "미완료", "완료율"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    for row, (cat, data) in enumerate(sorted(categories.items()), 4):
        incomplete = data["total"] - data["completed"]
        rate = f'{data["completed"] / data["total"] * 100:.1f}%' if data["total"] > 0 else "0%"

        for col, value in enumerate([cat, data["total"], data["completed"], incomplete, rate], 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    for col, width in enumerate([20, 10, 10, 10, 12], 1):
        ws.column_dimensions[chr(64 + col)].width = width


async def import_reminders_from_excel(
    db: AsyncSession,
    company_id: UUID,
    user_id: UUID,
    file: UploadFile,
) -> list[dict]:
    """Excel 파일에서 리마인더를 가져옵니다."""
    member_result = await db.execute(
        select(CompanyMember).where(
            CompanyMember.user_id == user_id,
            CompanyMember.company_id == company_id,
        )
    )
    if not member_result.scalar_one_or_none():
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty file")

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Excel file format. Please upload a valid .xlsx file.",
        )
    ws = wb.active

    imported = []
    errors = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row_idx, row in enumerate(rows, 2):
        if not row or not row[0]:
            continue

        try:
            title = str(row[1]) if len(row) > 1 and row[1] else None
            category = str(row[2]) if len(row) > 2 and row[2] else None
            deadline_str = str(row[3]) if len(row) > 3 and row[3] else None

            if not title or not category or not deadline_str:
                errors.append({"row": row_idx, "error": "필수 필드 누락 (제목, 카테고리, 마감일)"})
                continue

            if isinstance(row[3], datetime):
                deadline = row[3].date()
            elif isinstance(row[3], date):
                deadline = row[3]
            else:
                deadline = datetime.strptime(deadline_str, "%Y-%m-%d").date()

            description = str(row[7]) if len(row) > 7 and row[7] else None

            reminder = Reminder(
                company_id=company_id,
                title=title,
                category=category,
                deadline=deadline,
                description=description,
                priority=0,
                created_by=user_id,
            )
            db.add(reminder)
            imported.append({
                "title": title,
                "category": category,
                "deadline": deadline.isoformat(),
            })

        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})

    await db.flush()

    return {"imported": imported, "imported_count": len(imported), "errors": errors}
